import io
from pdf2image import convert_from_path
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as ExcelImage
import app
import base64
import re
import os
from dotenv import load_dotenv
import ext_samples

load_dotenv()

poppler_Path = os.getenv("poppler_path")

prompt = """
Guideines: 
1. Start Alt-text with the type of image "An illustration" or "A diagram " or "A flowchart ".
2. Describe all visible elements, including labels, symbols, and numbers.
3. Use space between abbreviations. For example, DNA should be D N A.
4. Do NOT include caption text or interpretation.
5. Short Alt Text: Maximum 200 characters.
6. Long Alt Text: No character limit. Use full sentences and include every necessary detail.
7. output format: Short Alt text \n\n Long Alt text
8. Do not write anything except the output format

Output Samples are given below....
"""


def check_uploaded():
    pass


def pdf_images_to_excel(pdf_path, excel_path, completed_pages, row_number, sample_file):
    #sample = ext_examples.extract_alt_text_examples(sample_file,num_examples=5)
    samples = ext_samples.extract_alt_text_examples(sample_file)
    # Convert PDF pages to images
    images = convert_from_path(pdf_path)
    # Create a new Excel workbook and sheet
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()

    ws = wb.active
    ws.title = "PDF Images"
    ws[f'A1'] = "Page No."
    ws[f'C1'] = "Figure"
    ws[f'E1'] = "Caption"
    ws[f'G1'] = "Short ALT-Text"
    ws[f'I1'] = "Long Alt-text"

    try:
        i=row_number
        for j, image in enumerate(images):
            if j<completed_pages:
                continue 
            # Convert image to bytes
            img_buffer = io.BytesIO()
            image.save(img_buffer, format="PNG")
            img_buffer.seek(0)

            # Convert to Base64
            img_base64 = base64.b64encode(img_buffer.read()).decode("utf-8") 

            img = ext_samples.crop_image(image)
            alt_text = app.generate_alt_text(samples, img, prompt) 

            data = alt_text.split('\n\n')
            short_text = (data[0].strip()).replace("Short alt text:", "")
            long_text = (data[1].strip()).replace("Long alt text:", "")

            print(short_text, "\n", long_text)
              
            ws[f'A{i+2}'] = f"Page no. {j+1}"
            ws[f'G{i+2}'] = short_text
            ws[f'I{i+2}'] = long_text
            i+=1
            print(f"page {j} completed")
        wb.save(excel_path)
        print(f"Excel saved at: {excel_path}")       
    except:
        wb.save(excel_path)
        print(f"Some error occured")
        print(f"Excel saved at: {excel_path}")


     
def main():
    completed_pages = int(input("Enter the number of completed pages: "))
    sapmle_path = "sample.xlsx"
    row_number = int(input("Enter the number of rows completed: "))
    pdf_images_to_excel("input.pdf", "output.xlsx", completed_pages, row_number, sapmle_path)


if __name__ == "__main__":
    main()